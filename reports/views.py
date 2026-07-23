from rest_framework import permissions
from rest_framework.views import APIView
from rest_framework.response import Response
from django.contrib.auth import get_user_model
from django.utils import timezone
from django.http import HttpResponse
from leaves.models import LeaveRequest
import calendar
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

User = get_user_model()


class TeamCalendarView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        import calendar as cal_module
        from datetime import date, timedelta
        from leaves.models import LeaveRequest
        from attendance.models import AttendanceSession

        year = int(request.query_params.get('year', timezone.now().year))
        month = int(request.query_params.get('month', timezone.now().month))
        dept_id = request.query_params.get('department')

        user = request.user

        if user.effective_role == 'admin':
            users = User.objects.filter(is_active=True)
            if dept_id:
                users = users.filter(department_id=dept_id)
        elif user.effective_role == 'director':
            users = User.objects.filter(is_active=True)
            if dept_id:
                users = users.filter(department_id=dept_id)
        elif user.effective_role == 'manager':
            users = User.objects.filter(is_active=True, department=user.department)
        else:
            users = User.objects.filter(id=user.id)

        num_days = cal_module.monthrange(year, month)[1]
        today = timezone.localdate()

        result = []
        for u in users:
            sessions = AttendanceSession.objects.filter(
                user=u, date__year=year, date__month=month, status__in=['complete', 'open'],
            )
            attended_days = set(s.date.day for s in sessions)

            leaves = LeaveRequest.objects.filter(
                user=u, status='approved',
            ).filter(
                start_date__lte=date(year, month, num_days),
                end_date__gte=date(year, month, 1),
            ).select_related('leave_type')

            leave_days = {}
            for leave in leaves:
                current = leave.start_date
                while current <= leave.end_date:
                    if current.year == year and current.month == month:
                        leave_days[current.day] = {
                            'leave_type': leave.leave_type.name,
                            'color': leave.leave_type.color,
                        }
                    current += timedelta(days=1)

            for day in range(1, num_days + 1):
                d = date(year, month, day)
                weekday = d.weekday()
                is_weekend = weekday >= 5
                if is_weekend:
                    continue
                status = None
                leave_type = None
                color = None
                if day in leave_days:
                    status = 'leave'
                    leave_type = leave_days[day]['leave_type']
                    color = leave_days[day]['color']
                elif day in attended_days:
                    status = 'present'
                    color = '#22C55E'
                elif d <= today:
                    status = 'absent'
                    color = '#EF4444'
                if status:
                    result.append({
                        'user_id': u.id,
                        'username': u.username,
                        'full_name': u.get_full_name() or u.username,
                        'department_id': u.department_id,
                        'department_name': u.department.name if u.department else None,
                        'date': str(d),
                        'status': status,
                        'leave_type': leave_type,
                        'color': color,
                    })

        return Response(result)


class AttendanceExportView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        year = int(request.query_params.get('year', timezone.now().year))
        month = int(request.query_params.get('month', timezone.now().month))
        user_id = request.query_params.get('user_id')

        user = request.user

        if user_id and user.role in ['admin', 'manager']:
            try:
                target_user = User.objects.get(id=user_id)
            except User.DoesNotExist:
                return Response({'detail': 'User not found.'}, status=404)
        else:
            target_user = user

        from attendance.models import AttendanceSession
        attendances = AttendanceSession.objects.filter(
            user=target_user, date__year=year, date__month=month, status='complete'
        ).order_by('date')

        wb = openpyxl.Workbook()
        ws = wb.active
        month_name = calendar.month_name[month]
        ws.title = f'{month_name} {year}'

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
        center = Alignment(horizontal='center', vertical='center')
        thin = Side(style='thin', color='D1D5DB')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        present_fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
        absent_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
        weekend_fill = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')

        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = f'Attendance Report — {target_user.get_full_name() or target_user.username} — {month_name} {year}'
        title_cell.font = Font(bold=True, size=13)
        title_cell.alignment = center
        ws.row_dimensions[1].height = 30

        headers = ['Date', 'Day', 'Check In', 'Check Out', 'Work Hours', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border
        ws.row_dimensions[2].height = 20

        num_days = calendar.monthrange(year, month)[1]
        attendance_map = {a.date.day: a for a in attendances}

        from datetime import date
        day_names = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
        total_hours = 0
        present_days = 0

        for day in range(1, num_days + 1):
            d = date(year, month, day)
            row = day + 2
            is_weekend = d.weekday() >= 5
            if day in attendance_map:
                a = attendance_map[day]
                check_in = a.clock_in.strftime('%H:%M') if a.clock_in else '-'
                check_out = a.clock_out.strftime('%H:%M') if a.clock_out else '-'
                work_hours = float(a.work_hours) if a.work_hours else 0
                status = 'Present'
                fill = present_fill
                if work_hours:
                    total_hours += work_hours
                    present_days += 1
            else:
                check_in = '-'
                check_out = '-'
                work_hours = '-'
                status = 'Weekend' if is_weekend else 'Absent'
                fill = weekend_fill if is_weekend else absent_fill

            row_data = [
                d.strftime('%d.%m.%Y'), day_names[d.weekday()],
                check_in, check_out,
                work_hours if work_hours != '-' else '-', status,
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.fill = fill
                cell.alignment = center
                cell.border = border

        summary_row = num_days + 4
        ws.cell(row=summary_row, column=1, value='Summary').font = Font(bold=True)
        ws.cell(row=summary_row + 1, column=1, value='Present Days:')
        ws.cell(row=summary_row + 1, column=2, value=present_days)
        ws.cell(row=summary_row + 2, column=1, value='Total Hours:')
        ws.cell(row=summary_row + 2, column=2, value=round(total_hours, 2))

        col_widths = [14, 12, 12, 12, 12, 12]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="attendance_{target_user.username}_{year}_{month:02d}.xlsx"'
        wb.save(response)
        return response


class LeaveExportView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from io import BytesIO

        year = int(request.query_params.get('year', timezone.now().year))
        user_id = request.query_params.get('user_id')
        user = request.user

        if user_id and user.role in ['admin', 'manager']:
            try:
                target_user = User.objects.get(id=user_id)
            except User.DoesNotExist:
                return Response({'detail': 'User not found.'}, status=404)
        else:
            target_user = user

        leaves = LeaveRequest.objects.filter(
            user=target_user, start_date__year=year
        ).select_related('leave_type').order_by('start_date')

        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer, pagesize=A4,
            rightMargin=2*cm, leftMargin=2*cm,
            topMargin=2*cm, bottomMargin=2*cm
        )

        styles = getSampleStyleSheet()
        elements = []

        title_style = ParagraphStyle(
            'Title', parent=styles['Title'], fontSize=16,
            spaceAfter=6, textColor=colors.HexColor('#1E40AF'),
        )
        subtitle_style = ParagraphStyle(
            'Subtitle', parent=styles['Normal'], fontSize=11,
            spaceAfter=20, textColor=colors.HexColor('#6B7280'),
        )

        elements.append(Paragraph('Leave Summary Report', title_style))
        elements.append(Paragraph(
            f'{target_user.get_full_name() or target_user.username} — {year}',
            subtitle_style
        ))

        table_data = [['Leave Type', 'Start Date', 'End Date', 'Days', 'Status', 'Reviewed By']]
        status_colors = {
            'approved': colors.HexColor('#D1FAE5'),
            'rejected': colors.HexColor('#FEE2E2'),
            'pending':  colors.HexColor('#FEF3C7'),
            'cancelled': colors.HexColor('#F3F4F6'),
        }

        row_styles = []
        for i, leave in enumerate(leaves, 1):
            table_data.append([
                leave.leave_type.name,
                leave.start_date.strftime('%d.%m.%Y'),
                leave.end_date.strftime('%d.%m.%Y'),
                str(int(leave.total_days)),
                leave.status.capitalize(),
                leave.reviewed_by.get_full_name() if leave.reviewed_by else '-',
            ])
            bg_color = status_colors.get(leave.status, colors.white)
            row_styles.append(('BACKGROUND', (0, i), (-1, i), bg_color))

        col_widths = [4.5*cm, 3*cm, 3*cm, 2*cm, 2.5*cm, 3.5*cm]
        table = Table(table_data, colWidths=col_widths)
        base_style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E40AF')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWHEIGHT', (0, 0), (-1, -1), 22),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D1D5DB')),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ] + row_styles
        table.setStyle(TableStyle(base_style))
        elements.append(table)
        elements.append(Spacer(1, 0.8*cm))

        elements.append(Paragraph('Summary by Leave Type', ParagraphStyle(
            'SectionTitle', parent=styles['Heading2'], fontSize=12,
            textColor=colors.HexColor('#1E40AF'), spaceBefore=10, spaceAfter=8,
        )))

        from leaves.models import LeaveBalance
        balances = LeaveBalance.objects.filter(
            user=target_user, year=year
        ).select_related('leave_type')

        summary_data = [['Leave Type', 'Total Days', 'Used Days', 'Remaining']]
        for balance in balances:
            summary_data.append([
                balance.leave_type.name,
                str(balance.total_days),
                str(int(balance.used_days)),
                str(int(balance.remaining_days)),
            ])

        if len(summary_data) > 1:
            summary_table = Table(summary_data, colWidths=[5*cm, 3*cm, 3*cm, 3*cm])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E40AF')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ROWHEIGHT', (0, 0), (-1, -1), 22),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D1D5DB')),
            ]))
            elements.append(summary_table)

        doc.build(elements)
        buffer.seek(0)

        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="leaves_{target_user.username}_{year}.pdf"'
        return response


class AdminDashboardView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if request.user.effective_role != 'admin':
            return Response({'detail': 'Permission denied.'}, status=403)

        from decimal import Decimal
        from django.db.models import Sum
        from attendance.models import AttendanceSession
        from accounts.models import Department

        today = timezone.localdate()
        current_year = today.year
        current_month = today.month

        active_employees = User.objects.filter(is_active=True, is_superuser=False)
        total_employees = active_employees.count()

        present_today = active_employees.filter(
            sessions__date=today, sessions__status__in=['complete', 'open'],
        ).distinct().count()
        on_leave_today = LeaveRequest.objects.filter(
            status='approved', start_date__lte=today, end_date__gte=today, user__in=active_employees,
        ).count()
        pending_leave_requests = LeaveRequest.objects.filter(status='pending').count()
        absent_today = max(0, total_employees - present_today - on_leave_today)

        month_hours = AttendanceSession.objects.filter(
            user__in=active_employees, date__year=current_year, date__month=current_month, status='complete',
        ).aggregate(s=Sum('work_hours'))['s'] or Decimal('0')
        avg_hours_this_month = round(float(month_hours) / total_employees, 1) if total_employees else 0

        departments = Department.objects.all()
        dept_data = [{
            'name': dept.name,
            'employee_count': User.objects.filter(department=dept, is_active=True, is_superuser=False).count(),
        } for dept in departments]

        today_sessions = AttendanceSession.objects.filter(
            date=today, user__in=active_employees,
        ).select_related('user', 'user__department').order_by('user_id', 'clock_in')
        rows_by_user = {}
        for s in today_sessions:
            row = rows_by_user.setdefault(s.user_id, {
                'full_name': s.user.get_full_name() or s.user.username,
                'department_name': s.user.department.name if s.user.department else None,
                'check_in': None, 'check_out': None, 'hours_worked': Decimal('0'),
            })
            if row['check_in'] is None:
                row['check_in'] = s.clock_in
            if s.clock_out:
                row['check_out'] = s.clock_out
            row['hours_worked'] += s.work_hours or Decimal('0')
        today_attendance = [{
            'full_name': r['full_name'],
            'department_name': r['department_name'],
            'check_in': r['check_in'].isoformat() if r['check_in'] else None,
            'check_out': r['check_out'].isoformat() if r['check_out'] else None,
            'hours_worked': float(r['hours_worked']),
            'status': 'present',
        } for r in rows_by_user.values()]

        recent_leaves = LeaveRequest.objects.select_related('user', 'leave_type').order_by('-created_at')[:8]
        recent_leave_requests = [{
            'id': leave.id,
            'employee_name': leave.user.get_full_name() or leave.user.username,
            'leave_type_name': leave.leave_type.name,
            'start_date': leave.start_date,
            'end_date': leave.end_date,
            'days_requested': float(leave.total_days),
            'status': leave.status,
        } for leave in recent_leaves]

        employees = [{
            'first_name': u.first_name,
            'last_name': u.last_name,
            'username': u.username,
            'full_name': u.get_full_name() or u.username,
            'role': u.effective_role,
            'department_name': u.department.name if u.department else None,
            'position': u.position,
            'is_active': u.is_active,
        } for u in User.objects.filter(is_superuser=False).select_related('department').order_by('last_name', 'first_name')]

        return Response({
            'total_employees': total_employees,
            'present_today': present_today,
            'on_leave_today': on_leave_today,
            'absent_today': absent_today,
            'avg_hours_this_month': avg_hours_this_month,
            'pending_leave_requests': pending_leave_requests,
            'departments': dept_data,
            'today_attendance': today_attendance,
            'recent_leave_requests': recent_leave_requests,
            'employees': employees,
        })


class ManagerDashboardView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if request.user.effective_role not in ['admin', 'manager']:
            return Response({'detail': 'Permission denied.'}, status=403)

        from decimal import Decimal
        from django.db.models import Sum
        from attendance.models import AttendanceSession

        user = request.user
        today = timezone.localdate()
        current_year = today.year
        current_month = today.month

        if user.effective_role == 'admin':
            team = User.objects.filter(is_active=True).select_related('department')
        else:
            team = User.objects.filter(is_active=True, department=user.department).select_related('department')
        total_team = team.count()

        present_today = team.filter(
            sessions__date=today, sessions__status__in=['complete', 'open'],
        ).distinct().count()
        on_leave_today = LeaveRequest.objects.filter(
            status='approved', start_date__lte=today, end_date__gte=today, user__in=team,
        ).count()
        pending_leaves = LeaveRequest.objects.filter(status='pending', user__in=team).count()

        month_hours = AttendanceSession.objects.filter(
            user__in=team, date__year=current_year, date__month=current_month, status='complete',
        ).aggregate(s=Sum('work_hours'))['s'] or Decimal('0')
        avg_hours_this_month = round(float(month_hours) / total_team, 1) if total_team else 0

        recent_leaves = LeaveRequest.objects.filter(
            status='pending', user__in=team
        ).select_related('user', 'leave_type').order_by('-created_at')[:5]
        recent_leaves_data = [{
            'id': leave.id,
            'employee_name': leave.user.get_full_name() or leave.user.username,
            'leave_type_name': leave.leave_type.name,
            'start_date': leave.start_date,
            'end_date': leave.end_date,
            'days_requested': float(leave.total_days),
        } for leave in recent_leaves]

        today_sessions = AttendanceSession.objects.filter(
            date=today, user__in=team,
        ).order_by('user_id', 'clock_in')
        sessions_by_user = {}
        for s in today_sessions:
            row = sessions_by_user.setdefault(s.user_id, {
                'check_in': None, 'check_out': None, 'hours_worked': Decimal('0'),
            })
            if row['check_in'] is None:
                row['check_in'] = s.clock_in
            if s.clock_out:
                row['check_out'] = s.clock_out
            row['hours_worked'] += s.work_hours or Decimal('0')

        team_status = []
        for member in team:
            on_leave = LeaveRequest.objects.filter(
                user=member, status='approved',
                start_date__lte=today, end_date__gte=today,
            ).select_related('leave_type').first()
            session_row = sessions_by_user.get(member.id)

            if on_leave:
                status_val = 'leave'
                detail = on_leave.leave_type.name
            elif session_row:
                status_val = 'present'
                detail = session_row['check_in'].strftime('%H:%M') if session_row['check_in'] else '-'
            else:
                status_val = 'absent'
                detail = '-'

            team_status.append({
                'user_id': member.id,
                'full_name': member.get_full_name() or member.username,
                'role': member.role,
                'position': member.position,
                'employee_number': member.employee_number,
                'department_name': member.department.name if member.department_id else None,
                'email': member.email,
                'status': status_val,
                'detail': detail,
                'check_in': session_row['check_in'].isoformat() if session_row and session_row['check_in'] else None,
                'check_out': session_row['check_out'].isoformat() if session_row and session_row['check_out'] else None,
                'hours_worked': float(session_row['hours_worked']) if session_row else None,
            })

        return Response({
            'stats': {
                'total_team': total_team,
                'present_today': present_today,
                'on_leave_today': on_leave_today,
                'pending_leaves': pending_leaves,
            },
            'present_today': present_today,
            'avg_hours_this_month': avg_hours_this_month,
            'team_status': team_status,
            'recent_pending_leaves': recent_leaves_data,
        })


class EmployeeDashboardView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        from decimal import Decimal
        from django.db.models import Sum
        from attendance.models import AttendanceSession
        from leaves.models import LeaveBalance, LeaveType

        today = timezone.localdate()
        current_year = today.year
        current_month = today.month
        user = request.user

        hours_this_month = AttendanceSession.objects.filter(
            user=user, date__year=current_year, date__month=current_month, status='complete',
        ).aggregate(s=Sum('work_hours'))['s'] or Decimal('0')

        def remaining_for(**lookup):
            leave_type = LeaveType.objects.filter(is_active=True, **lookup).first()
            if not leave_type:
                return None
            balance = LeaveBalance.objects.filter(user=user, leave_type=leave_type, year=current_year).first()
            return float(balance.remaining_days) if balance else None

        leave_balance = {
            'annual': remaining_for(name='Annual Leave'),
            'sick': remaining_for(is_sick_leave=True),
        }

        pending_leave_requests = LeaveRequest.objects.filter(user=user, status='pending').count()

        recent_dates = list(
            AttendanceSession.objects.filter(user=user).order_by('-date')
            .values_list('date', flat=True).distinct()[:10]
        )
        recent_sessions = AttendanceSession.objects.filter(
            user=user, date__in=recent_dates
        ).order_by('date', 'clock_in')
        days_map = {}
        for s in recent_sessions:
            entry = days_map.setdefault(s.date, {'check_in': None, 'check_out': None, 'hours_worked': Decimal('0')})
            if entry['check_in'] is None:
                entry['check_in'] = s.clock_in
            if s.clock_out:
                entry['check_out'] = s.clock_out
            entry['hours_worked'] += s.work_hours or Decimal('0')
        recent_attendance = [{
            'date': str(d),
            'check_in': days_map[d]['check_in'].isoformat() if days_map[d]['check_in'] else None,
            'check_out': days_map[d]['check_out'].isoformat() if days_map[d]['check_out'] else None,
            'hours_worked': float(days_map[d]['hours_worked']),
        } for d in sorted(days_map.keys(), reverse=True)]

        return Response({
            'hours_this_month': float(hours_this_month),
            'leave_balance': leave_balance,
            'pending_leave_requests': pending_leave_requests,
            'recent_attendance': recent_attendance,
        })


class PontajExportView(APIView):
    """Exportă exact datele salvate în PontajSheet/PontajEntry (aceleași
    afișate/editate în pagina /pontaj) — nu recalculează pe loc din sesiuni
    de lucru, ca să reflecte editările manuale și starea aprobată a foii."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        from collections import defaultdict
        from datetime import date

        from accounts.models import Department
        from leaves.utils import get_public_holidays_named

        from .pontaj import get_or_create_personal_sheet, get_or_create_sheet, sync_leave_requests

        year = int(request.query_params.get('year', timezone.now().year))
        month = int(request.query_params.get('month', timezone.now().month))
        department_id = request.query_params.get('department_id')
        user_id = request.query_params.get('user_id')
        user = request.user

        scope_department = None
        if user_id:
            if user.effective_role not in ['admin', 'manager', 'director']:
                return Response({'detail': 'Permission denied.'}, status=403)
            users = User.objects.filter(id=user_id, is_active=True).select_related('department__schedule_type')
        elif department_id:
            if user.effective_role not in ['admin', 'manager', 'director']:
                return Response({'detail': 'Permission denied.'}, status=403)
            scope_department = Department.objects.filter(id=department_id).first()
            users = User.objects.filter(
                department_id=department_id, is_active=True
            ).select_related('department__schedule_type').order_by('last_name', 'first_name')
        elif user.effective_role == 'manager':
            scope_department = user.department
            users = User.objects.filter(
                department=user.department, is_active=True
            ).select_related('department__schedule_type').order_by('last_name', 'first_name')
        elif user.effective_role in ['admin', 'director']:
            users = User.objects.filter(is_active=True).select_related('department__schedule_type').order_by('last_name', 'first_name')
        else:
            users = User.objects.filter(id=user.id).select_related('department__schedule_type')

        users_list = list(users)
        if user_id and users_list:
            scope_department = users_list[0].department

        num_days = calendar.monthrange(year, month)[1]
        RO_MONTHS = [
            '', 'ianuarie', 'februarie', 'martie', 'aprilie', 'mai', 'iunie',
            'iulie', 'august', 'septembrie', 'octombrie', 'noiembrie', 'decembrie',
        ]
        month_name_ro = RO_MONTHS[month]

        if scope_department:
            scope_label = scope_department.name
            manager = User.objects.filter(
                department=scope_department, role='manager', is_active=True
            ).order_by('last_name', 'first_name').first()
        else:
            scope_label = users_list[0].get_full_name() if user_id and users_list else 'Toate departamentele'
            manager = None
        if not manager:
            manager = User.objects.filter(role='director', is_active=True).order_by('last_name', 'first_name').first()
        manager_name = manager.get_full_name() if manager else ''

        # Sursa reala a datelor: foile de pontaj salvate (departament sau
        # personale), nu o recalculare live din sesiuni/concedii.
        entries_by_user = {}
        dept_ids = {u.department_id for u in users_list if u.department_id}
        for dept in Department.objects.filter(id__in=dept_ids):
            sheet = get_or_create_sheet(dept, year, month)
            sync_leave_requests(sheet)
            per_day = defaultdict(dict)
            for entry in sheet.entries.all():
                per_day[entry.user_id][entry.day] = entry
            entries_by_user.update(per_day)
        for u in users_list:
            if u.department_id is None and u.id not in entries_by_user:
                sheet = get_or_create_personal_sheet(u, year, month)
                sync_leave_requests(sheet)
                per_day = defaultdict(dict)
                for entry in sheet.entries.all():
                    per_day[entry.user_id][entry.day] = entry
                entries_by_user.update(per_day)

        holidays = {d.day: name for d, name in get_public_holidays_named(year).items() if d.month == month}
        norm_hours = sum(
            8 for day in range(1, num_days + 1)
            if date(year, month, day).weekday() < 5 and day not in holidays
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f'Pontaj {month_name_ro} {year}'[:31]

        thin = Side(style='thin', color='000000')
        border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill   = PatternFill('solid', start_color='D9E1F2', end_color='D9E1F2')
        weekend_fill  = PatternFill('solid', start_color='D9D9D9', end_color='D9D9D9')
        holiday_fill  = PatternFill('solid', start_color='FCE4EC', end_color='FCE4EC')
        present_fill  = PatternFill('solid', start_color='E2EFDA', end_color='E2EFDA')
        leave_co_fill = PatternFill('solid', start_color='FFF2CC', end_color='FFF2CC')
        leave_cm_fill = PatternFill('solid', start_color='FCE4D6', end_color='FCE4D6')
        leave_other_fill = PatternFill('solid', start_color='EDD9FC', end_color='EDD9FC')

        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left   = Alignment(horizontal='left', vertical='center', wrap_text=True)
        bold   = Font(bold=True, size=9, name='Arial')
        normal = Font(size=9, name='Arial')
        small  = Font(size=8, name='Arial')

        LEAVE_CODES = ('CO', 'CM', 'FP', 'IC', 'CI', 'AC', 'NE')
        total_cols = [
            ('Norma', 'T_NORM'),
            ('Total\nOre', 'T_ORE'),
            *((code, f'T_{code}') for code in LEAVE_CODES),
            ('Zile\nnelucrate', 'T_UNWORKED'),
        ]

        # ── Header companie ──────────────────────────────────────────────────
        last_col = get_column_letter(4 + num_days + len(total_cols))
        ws.merge_cells(f'A1:{last_col}1')
        ws['A1'] = f'Nexoria Group — Fișă lunară de pontaj — {scope_label} — {month_name_ro} {year}'
        ws['A1'].alignment = center
        ws['A1'].fill = PatternFill('solid', start_color='1E3A5F', end_color='1E3A5F')
        ws['A1'].font = Font(bold=True, size=11, name='Arial', color='FFFFFF')
        ws.row_dimensions[1].height = 22

        # ── Antet coloane ────────────────────────────────────────────────────
        ws.merge_cells('A2:A3')
        ws['A2'] = 'No.'
        ws['A2'].font = bold
        ws['A2'].alignment = center
        ws['A2'].border = border_all
        ws['A2'].fill = header_fill

        ws.merge_cells('B2:B3')
        ws['B2'] = 'Full Name'
        ws['B2'].font = bold
        ws['B2'].alignment = center
        ws['B2'].border = border_all
        ws['B2'].fill = header_fill

        ws.merge_cells('C2:C3')
        ws['C2'] = 'Funcție'
        ws['C2'].font = bold
        ws['C2'].alignment = center
        ws['C2'].border = border_all
        ws['C2'].fill = header_fill

        ws.merge_cells('D2:D3')
        ws['D2'] = 'Marcă'
        ws['D2'].font = bold
        ws['D2'].alignment = center
        ws['D2'].border = border_all
        ws['D2'].fill = header_fill

        day_names_en = ['SU', 'MO', 'TU', 'WE', 'TH', 'FR', 'SA']
        start_col = 5

        for day in range(1, num_days + 1):
            d = date(year, month, day)
            col = start_col + day - 1
            is_holiday = day in holidays

            cell1 = ws.cell(row=2, column=col, value=day)
            cell1.font = bold
            cell1.alignment = center
            cell1.border = border_all

            cell2 = ws.cell(row=3, column=col, value=day_names_en[d.weekday()])
            cell2.font = small
            cell2.alignment = center
            cell2.border = border_all

            if d.weekday() >= 5:
                cell1.fill = weekend_fill
                cell2.fill = weekend_fill
            elif is_holiday:
                cell1.fill = holiday_fill
                cell2.fill = holiday_fill
                cell1.comment = Comment(holidays[day], 'HCM487')
            else:
                cell1.fill = header_fill
                cell2.fill = header_fill

        total_start_col = start_col + num_days
        for i, (label, key) in enumerate(total_cols):
            col = total_start_col + i
            ws.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)
            cell = ws.cell(row=2, column=col, value=label)
            cell.font = bold
            cell.alignment = center
            cell.fill = header_fill
            cell.border = border_all

        # ── Date angajați ─────────────────────────────────────────────────────
        data_start_row = 4

        for idx, u in enumerate(users_list):
            row = data_start_row + idx
            day_entries = entries_by_user.get(u.id, {})

            cell_nr = ws.cell(row=row, column=1, value=idx + 1)
            cell_nr.font = normal
            cell_nr.alignment = center
            cell_nr.border = border_all

            cell_name = ws.cell(row=row, column=2, value=u.get_full_name() or u.username)
            cell_name.font = normal
            cell_name.alignment = left
            cell_name.border = border_all

            cell_position = ws.cell(row=row, column=3, value=u.position or '')
            cell_position.font = normal
            cell_position.alignment = left
            cell_position.border = border_all

            cell_marca = ws.cell(row=row, column=4, value=u.employee_number or '')
            cell_marca.font = normal
            cell_marca.alignment = center
            cell_marca.border = border_all

            worked_hours = 0.0
            leave_counts = {code: 0 for code in LEAVE_CODES}

            for day in range(1, num_days + 1):
                d = date(year, month, day)
                col = start_col + day - 1
                cell = ws.cell(row=row, column=col)
                cell.font = normal
                cell.alignment = center
                cell.border = border_all

                if d.weekday() >= 5:
                    cell.fill = weekend_fill
                    continue

                entry = day_entries.get(day)
                code = entry.leave_code if entry else ''
                hours = float(entry.hours) if entry and entry.hours is not None else None

                if code in leave_counts:
                    cell.value = code
                    cell.fill = leave_co_fill if code == 'CO' else leave_cm_fill if code == 'CM' else leave_other_fill
                    leave_counts[code] += 1
                elif hours is not None:
                    cell.value = hours
                    cell.fill = present_fill
                    worked_hours += hours
                elif day in holidays:
                    cell.fill = holiday_fill

            leave_hours = sum(leave_counts.values()) * 8
            total_hours = round(worked_hours + leave_hours, 1)
            row_values = [norm_hours, total_hours, *(leave_counts[c] for c in LEAVE_CODES), leave_hours]

            for i, val in enumerate(row_values):
                col = total_start_col + i
                tc = ws.cell(row=row, column=col, value=val)
                tc.font = bold if val else normal
                tc.alignment = center
                tc.border = border_all
                tc.fill = header_fill

        # ── Lățimi coloane ────────────────────────────────────────────────────
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 24
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 10
        for day in range(1, num_days + 1):
            ws.column_dimensions[get_column_letter(start_col + day - 1)].width = 3.5
        for i in range(len(total_cols)):
            ws.column_dimensions[get_column_letter(total_start_col + i)].width = 7

        ws.row_dimensions[2].height = 18
        ws.row_dimensions[3].height = 14
        for idx in range(len(users_list)):
            ws.row_dimensions[data_start_row + idx].height = 16

        ws.freeze_panes = f'{get_column_letter(start_col)}4'

        # ── Manager ───────────────────────────────────────────────────────────
        if manager_name:
            manager_row = data_start_row + len(users_list) + 1
            ws.cell(row=manager_row, column=total_start_col, value='Manager:').font = bold
            ws.merge_cells(
                start_row=manager_row, start_column=total_start_col + 1,
                end_row=manager_row, end_column=total_start_col + len(total_cols) - 1,
            )
            name_cell = ws.cell(row=manager_row, column=total_start_col + 1, value=manager_name)
            name_cell.font = normal
            name_cell.alignment = left

        # ── Legendă ───────────────────────────────────────────────────────────
        legend_row = data_start_row + len(users_list) + 2
        legend = [
            ('CO', 'Annual Leave'),
            ('CM', 'Sick Leave'),
            ('FP', 'Unpaid Leave'),
            ('IC', 'Parental / Child Care Leave'),
            ('CI', 'Carer Leave'),
            ('AC', 'Other paid leave'),
            ('NE', 'Unexcused absence'),
        ]
        ws.cell(row=legend_row, column=1, value='Legend:').font = bold
        for i, (code, desc) in enumerate(legend):
            ws.cell(row=legend_row + i + 1, column=1, value=code).font = bold
            ws.cell(row=legend_row + i + 1, column=2, value=desc).font = normal

        if holidays:
            holiday_legend_row = legend_row + len(legend) + 2
            ws.cell(row=holiday_legend_row, column=1, value='Public Holidays:').font = bold
            for i, day in enumerate(sorted(holidays)):
                ws.cell(row=holiday_legend_row + i + 1, column=1, value=f'{day:02d}.{month:02d}.{year}').font = normal
                ws.cell(row=holiday_legend_row + i + 1, column=2, value=holidays[day]).font = normal

        # ── Response ──────────────────────────────────────────────────────────
        dept_label = ''
        if scope_department:
            dept_label = f'_{scope_department.name.replace(" ", "_")}'
        elif user_id and users_list:
            dept_label = f'_{users_list[0].username}'

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'attendance_sheet{dept_label}_{year}_{month:02d}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


class DebugTeamView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        user = request.user
        team = User.objects.filter(manager=user)
        return Response({
            'current_user': user.username,
            'current_user_id': user.id,
            'team_count': team.count(),
            'team': list(team.values('username', 'manager_id')),
        })